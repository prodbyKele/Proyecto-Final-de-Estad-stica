import openpyxl #pip install openpyxl
import pandas as pd #pip install pandas
import matplotlib.pyplot as plt #pip install matplotlib
from matplotlib.ticker import PercentFormatter
from tabulate import tabulate #pip install tabulate

#basepath =  r"C:\Users\Kelet\iCloudDrive\UMG\Cuarto Semestre\Estadística"
#archivo = basepath + r"\Distribución de Frecuencias Simples 1.2.xlsx"

def import_excel():
  excel_dataframe = openpyxl.load_workbook("Pruebas.xlsx")

  dataframe = excel_dataframe.active

  data = []
  
  filas = int(input("Ingrese hasta que fila quiere hacer los cálculos: ")) + 1
  
  for row in range (1, filas):
    
    new_row = [row, ]
    
    for col in dataframe.iter_cols(2, dataframe.max_column):
      new_row.append(col[row].value)
      
    data.append(new_row)
      
  headers = ["#", "Genero", "Edad"]
  headers_align = (("center",)*3)

  print(tabulate(data, headers=headers, tablefmt='fancy_grid', colalign=headers_align))
  
#import_excel()

def distribucion_de_frecuencias_simple():
  data = import_excel()
  
  #frecuencia absoluta de cada dato
  f= {} #frecuencia absoluta
  for dato in data:
    f[dato] = f.get(dato, 0) + 1
    
  #frecuencia relativa y el porcentaje de cada dato
  total_datos = len(data)
  fr = {}
  porcentaje = {}
  for dato, frecuencia in f.items():
    fr[dato] = frecuencia / total_datos
    porcentaje[dato] = fr[dato]*100
    
  #frecuencia absoluta acumulada
  fa = {}
  faporcentaje = {}
  acum = 0
  for dato, frecuencia in f.items():
    acum += frecuencia
    fa[dato] = acum
    faporcentaje[dato] = fa[dato]*100
    
  #frecuencia descendente
  fa = {}
  faporcentaje = {}
  acum = 0
  for dato, frecuencia in f.items():
    acum -= frecuencia
    fa[dato] = acum
    faporcentaje[dato] = fa[dato]*100
    
  
  
  print(data)
  
distribucion_de_frecuencias_simple()